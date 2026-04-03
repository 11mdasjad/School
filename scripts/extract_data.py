"""
Extract Royal International Public School data from Excel and generate TypeScript demo-data.
"""
import openpyxl
import json
import re

FILE = '/Users/mohammad/Desktop/RPS/Royal_International_Public_School_Complete_Data (1).xlsx'
wb = openpyxl.load_workbook(FILE)

# ====== EXTRACT SCHOOL INFO ======
school_info = wb['School Information']
school_name = school_info.cell(1, 1).value  # ROYAL INTERNATIONAL PUBLIC SCHOOL
school_address = school_info.cell(3, 2).value
school_phone = school_info.cell(4, 2).value
school_email = school_info.cell(5, 2).value
principal = school_info.cell(7, 2).value

# ====== EXTRACT CLASS SUMMARY (class teacher mapping) ======
cs = wb['Class Summary']
class_teacher_map = {}
for row in cs.iter_rows(min_row=2, max_row=14, values_only=True):
    if row[0]:
        class_teacher_map[str(row[0]).strip()] = {
            'sec_a': row[1],
            'sec_b': row[2],
            'total': row[3],
            'teacher': row[4],
        }

# ====== EXTRACT TEACHERS ======
tw = wb['Teachers']
teachers = []
for row in tw.iter_rows(min_row=2, max_row=tw.max_row, values_only=True):
    if not row[0]:
        continue
    teachers.append({
        'sno': row[0],
        'employee_id': str(row[1]).strip(),
        'name': str(row[2]).strip(),
        'gender': str(row[3]).strip(),
        'subject': str(row[4]).strip(),
        'qualification': str(row[5]).strip() if row[5] else '',
        'experience': row[6],
        'classes_assigned': str(row[7]).strip() if row[7] else '',
        'date_of_joining': str(row[8]).strip() if row[8] else '',
        'mobile': str(row[9]).strip() if row[9] else '',
        'personal_email': str(row[10]).strip() if row[10] else '',
        'school_email': str(row[11]).strip() if row[11] else '',
        'address': str(row[12]).strip() if row[12] else '',
    })

# ====== EXTRACT ALL STUDENTS ======
class_sheets = ['Nursery', 'LKG', 'UKG', 'Class_1', 'Class_2', 'Class_3', 'Class_4', 'Class_5', 'Class_6', 'Class_7', 'Class_8', 'Class_9', 'Class_10']

students = []
for sheet_name in class_sheets:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        students.append({
            'sno': row[0],
            'admission_no': str(row[1]).strip(),
            'name': str(row[2]).strip(),
            'class': str(row[3]).strip(),
            'section': str(row[4]).strip(),
            'gender': str(row[5]).strip().lower(),
            'dob': str(row[6]).strip() if row[6] else '',
            'blood_group': str(row[7]).strip() if row[7] else '',
            'father_name': str(row[8]).strip(),
            'mother_name': str(row[9]).strip(),
            'father_mobile': str(row[10]).strip() if row[10] else '',
            'mother_mobile': str(row[11]).strip() if row[11] else '',
            'emergency_contact': str(row[12]).strip() if row[12] else '',
            'email': str(row[13]).strip() if row[13] else '',
            'address': str(row[14]).strip() if row[14] else '',
            'aadhar_no': str(row[15]).strip() if row[15] else '',
            'fee_status': str(row[21]).strip() if row[21] else '',
        })

print(f"Extracted {len(students)} students, {len(teachers)} teachers")

# ====== BUILD CLASS/SECTION MAPPING ======
class_names_order = ['Nursery', 'LKG', 'UKG', 'Class 1', 'Class 2', 'Class 3', 'Class 4', 'Class 5', 'Class 6', 'Class 7', 'Class 8', 'Class 9', 'Class 10']
classes = []
sections = []
for i, cn in enumerate(class_names_order, 1):
    cls_id = f'c{i}'
    grade = str(i) if i > 3 else ('N' if i == 1 else 'LKG' if i == 2 else 'UKG')
    classes.append({'id': cls_id, 'name': cn, 'grade_level': grade})
    sections.append({'id': f's{i*2-1}', 'class_id': cls_id, 'name': 'A'})
    sections.append({'id': f's{i*2}', 'class_id': cls_id, 'name': 'B'})

class_id_map = {cn: f'c{i+1}' for i, cn in enumerate(class_names_order)}

def get_section_id(class_name, section):
    cls_idx = class_names_order.index(class_name)
    sec_offset = 0 if section.strip().upper() == 'A' else 1
    return f's{(cls_idx) * 2 + 1 + sec_offset}'

# ====== BUILD PARENT MAP (by father_mobile) ======
parent_map = {}  # phone -> {father_name, mother_name, ...}
for s in students:
    phone = s['father_mobile']
    if phone and phone not in parent_map:
        parent_map[phone] = {
            'father_name': s['father_name'],
            'mother_name': s['mother_name'],
            'father_mobile': s['father_mobile'],
            'mother_mobile': s['mother_mobile'],
            'address': s['address'],
        }

print(f"Unique parents: {len(parent_map)}")

# ====== GENERATE TYPESCRIPT ======

# --- Profiles ---
# Admin profile (principal)
profiles = []
profiles.append(f"  {{ id: 'u-admin', full_name: '{principal}', email: 'principal@rips.edu.in', role: 'admin' as const, phone: '9876543210', avatar_url: '', status: 'active' as const, created_at: '2024-01-01', updated_at: '2024-01-01' }},")

# Teacher profiles
for i, t in enumerate(teachers):
    tid = f'u-t{i+1}'
    name = t['name'].replace("'", "\\'")
    email = t['school_email'] or t['personal_email'] or ''
    phone = t['mobile']
    profiles.append(f"  {{ id: '{tid}', full_name: '{name}', email: '{email}', role: 'teacher' as const, phone: '{phone}', avatar_url: '', status: 'active' as const, created_at: '2024-01-01', updated_at: '2024-01-01' }},")

# Parent profiles (by father_mobile)
parent_ids = {}
for i, (phone, pdata) in enumerate(parent_map.items()):
    pid = f'u-p{i+1}'
    parent_ids[phone] = pid
    fname = pdata['father_name'].replace("'", "\\'")
    profiles.append(f"  {{ id: '{pid}', full_name: '{fname}', email: '', role: 'parent' as const, phone: '{phone}', avatar_url: '', status: 'active' as const, created_at: '2024-01-01', updated_at: '2024-01-01' }},")

# Student profiles
for i, s in enumerate(students):
    sid = f'u-s{i+1}'
    name = s['name'].replace("'", "\\'")
    email = s['email'] or ''
    phone = s['father_mobile']  # login with father's phone
    profiles.append(f"  {{ id: '{sid}', full_name: '{name}', email: '{email}', role: 'student' as const, phone: '{phone}', avatar_url: '', status: 'active' as const, created_at: '2024-01-01', updated_at: '2024-01-01' }},")

# --- Teachers data ---
teacher_records = []
for i, t in enumerate(teachers):
    tid = f't{i+1}'
    profile_id = f'u-t{i+1}'
    emp_id = t['employee_id']
    spec = t['subject'].replace("'", "\\'")
    teacher_records.append(f"  {{ id: '{tid}', profile_id: '{profile_id}', employee_id: '{emp_id}', specialization: '{spec}', created_at: '2024-01-01' }},")

# --- Parents data ---
parent_records = []
for i, (phone, pdata) in enumerate(parent_map.items()):
    pid = f'p{i+1}'
    profile_id = parent_ids[phone]
    parent_records.append(f"  {{ id: '{pid}', profile_id: '{profile_id}', relationship: 'Father', created_at: '2024-01-01' }},")

# --- Students data ---
student_records = []
for i, s in enumerate(students):
    sid = f'st{i+1}'
    profile_id = f'u-s{i+1}'
    adm = s['admission_no']
    roll = str(s['sno'])
    cls_name = s['class']
    cls_id = class_id_map.get(cls_name, 'c1')
    sec_id = get_section_id(cls_name, s['section'])
    p_phone = s['father_mobile']
    parent_rec_id = ''
    for pi, (phone, _) in enumerate(parent_map.items()):
        if phone == p_phone:
            parent_rec_id = f'p{pi+1}'
            break
    gender = s['gender'] if s['gender'] in ['male', 'female'] else 'other'
    dob = s['dob']
    addr = s['address'].replace("'", "\\'").replace('\n', ' ')
    student_records.append(f"  {{ id: '{sid}', profile_id: '{profile_id}', admission_no: '{adm}', roll_no: '{roll}', class_id: '{cls_id}', section_id: '{sec_id}', parent_id: '{parent_rec_id}', dob: '{dob}', gender: '{gender}', address: '{addr}', created_at: '2024-01-01' }},")

# --- Subjects ---
subjects_list = sorted(set(t['subject'] for t in teachers))
subject_records = []
for i, sub in enumerate(subjects_list):
    sub_id = f'sub{i+1}'
    code = ''.join(w[0] for w in sub.split()).upper()[:4]
    sub_clean = sub.replace("'", "\\'")
    subject_records.append(f"  {{ id: '{sub_id}', name: '{sub_clean}', code: '{code}', description: '', created_at: '2024-01-01' }},")

# --- Teacher Assignments ---
subject_id_map = {sub: f'sub{i+1}' for i, sub in enumerate(subjects_list)}
assignment_records = []
a_id = 1
for i, t in enumerate(teachers):
    tid = f't{i+1}'
    sub_id = subject_id_map.get(t['subject'], 'sub1')
    assigned = t['classes_assigned']
    if assigned:
        for cls_str in assigned.split(','):
            cls_str = cls_str.strip()
            if cls_str in class_id_map:
                cls_id = class_id_map[cls_str]
                # Assign to section A by default
                cls_idx = class_names_order.index(cls_str)
                sec_id = f's{cls_idx * 2 + 1}'
                assignment_records.append(f"  {{ id: 'ta{a_id}', teacher_id: '{tid}', class_id: '{cls_id}', section_id: '{sec_id}', subject_id: '{sub_id}', created_at: '2024-01-01' }},")
                a_id += 1

# --- DEMO_CREDENTIALS for constants ---
# First teacher phone, first student's father, first parent
admin_phone = '9876543210'
first_teacher_phone = teachers[0]['mobile'] if teachers else '9000000002'
# First student of a higher class (not nursery)
first_student_phone = students[0]['father_mobile'] if students else '9000000003'
first_parent_phone = list(parent_map.keys())[0] if parent_map else '9000000004'

print(f"\n=== DEMO CREDENTIALS ===")
print(f"Admin: {admin_phone}")
print(f"Teacher ({teachers[0]['name']}): {first_teacher_phone}")
print(f"Student ({students[0]['name']} via father): {first_student_phone}")
print(f"Parent ({parent_map[first_parent_phone]['father_name']}): {first_parent_phone}")

# --- Write output ---
# Save the credential info for later use
creds_data = {
    'admin': admin_phone,
    'teacher': first_teacher_phone,
    'teacher_name': teachers[0]['name'],
    'student': first_student_phone,
    'student_name': students[0]['name'],
    'parent': first_parent_phone,
    'parent_name': parent_map[first_parent_phone]['father_name'],
}
with open('/Users/mohammad/Desktop/RPS/smart-school-attendance/scripts/creds.json', 'w') as f:
    json.dump(creds_data, f, indent=2)

# --- Classes TS ---
classes_ts = []
for c in classes:
    classes_ts.append(f"  {{ id: '{c['id']}', name: '{c['name']}', grade_level: '{c['grade_level']}', academic_year_id: 'ay1', created_at: '2024-01-01' }},")

sections_ts = []
for s in sections:
    sections_ts.append(f"  {{ id: '{s['id']}', class_id: '{s['class_id']}', name: '{s['name']}', created_at: '2024-01-01' }},")

# --- School Settings ---
school_addr = (school_address or '').replace("'", "\\'")
school_ph = str(school_phone or '').replace("'", "\\'")

# ---------- WRITE THE BIG FILE ----------
output = f'''// ============================================================
// Royal International Public School - Complete Data
// Generated from Excel: Royal_International_Public_School_Complete_Data.xlsx
// Total: {len(students)} students, {len(teachers)} teachers, {len(parent_map)} parents
// ============================================================

import {{
  Profile, AcademicYear, Class, Section, Subject,
  Teacher, Parent, Student, TeacherAssignment,
  AttendanceRecord, LeaveRequest, Notification,
  Holiday, SchoolSettings, AttendanceStatus,
}} from '@/types';

// ---- Profiles ({len(profiles)} total) ----
export const demoProfiles: Profile[] = [
{chr(10).join(profiles)}
];

// ---- Academic Years ----
export const demoAcademicYears: AcademicYear[] = [
  {{ id: 'ay1', name: '2024-2025', start_date: '2024-04-01', end_date: '2025-03-31', is_active: true, created_at: '2024-01-01' }},
  {{ id: 'ay2', name: '2023-2024', start_date: '2023-04-01', end_date: '2024-03-31', is_active: false, created_at: '2023-01-01' }},
];

// ---- Classes ({len(classes)} total) ----
export const demoClasses: Class[] = [
{chr(10).join(classes_ts)}
];

// ---- Sections ({len(sections)} total) ----
export const demoSections: Section[] = [
{chr(10).join(sections_ts)}
];

// ---- Subjects ({len(subject_records)} total) ----
export const demoSubjects: Subject[] = [
{chr(10).join(subject_records)}
];

// ---- Teachers ({len(teacher_records)} total) ----
export const demoTeachers: Teacher[] = [
{chr(10).join(teacher_records)}
];

// ---- Parents ({len(parent_records)} total) ----
export const demoParents: Parent[] = [
{chr(10).join(parent_records)}
];

// ---- Students ({len(student_records)} total) ----
export const demoStudents: Student[] = [
{chr(10).join(student_records)}
];

// ---- Teacher Assignments ({len(assignment_records)} total) ----
export const demoAssignments: TeacherAssignment[] = [
{chr(10).join(assignment_records)}
];

// ---- Generate Attendance Records ----
function generateAttendance(): AttendanceRecord[] {{
  const records: AttendanceRecord[] = [];
  const statuses: AttendanceStatus[] = ['present', 'present', 'present', 'present', 'present', 'present', 'present', 'absent', 'late', 'excused'];
  // Only generate for a subset (first 50 students) for performance
  const subset = demoStudents.slice(0, 50);
  let id = 1;

  for (let day = 1; day <= 30; day++) {{
    const date = new Date(2025, 2, day);
    if (date.getDay() === 0 || date.getDay() === 6) continue;

    for (const student of subset) {{
      const status = statuses[Math.floor(Math.random() * statuses.length)];
      records.push({{
        id: `att${{id}}`,
        student_id: student.id,
        class_id: student.class_id,
        section_id: student.section_id,
        subject_id: 'sub1',
        attendance_date: date.toISOString().split('T')[0],
        period_no: 1,
        status,
        remarks: status === 'late' ? 'Arrived late' : status === 'absent' ? 'No information' : undefined,
        marked_by: 'u-t1',
        created_at: date.toISOString(),
        updated_at: date.toISOString(),
      }});
      id++;
    }}
  }}
  return records;
}}

export const demoAttendance: AttendanceRecord[] = generateAttendance();

// ---- Leave Requests ----
export const demoLeaveRequests: LeaveRequest[] = [
  {{ id: 'lr1', student_id: 'st1', requested_by: '{list(parent_ids.values())[0]}', from_date: '2025-03-10', to_date: '2025-03-12', reason: 'Family event', status: 'approved', approved_by: 'u-admin', approved_at: '2025-03-09', created_at: '2025-03-08' }},
  {{ id: 'lr2', student_id: 'st3', requested_by: '{list(parent_ids.values())[2] if len(parent_ids) > 2 else "u-p3"}', from_date: '2025-03-15', to_date: '2025-03-15', reason: 'Medical appointment', status: 'pending', created_at: '2025-03-14' }},
  {{ id: 'lr3', student_id: 'st5', requested_by: '{list(parent_ids.values())[4] if len(parent_ids) > 4 else "u-p5"}', from_date: '2025-03-20', to_date: '2025-03-22', reason: 'Family emergency', status: 'pending', created_at: '2025-03-19' }},
  {{ id: 'lr4', student_id: 'st10', requested_by: '{list(parent_ids.values())[9] if len(parent_ids) > 9 else "u-p10"}', from_date: '2025-03-25', to_date: '2025-03-28', reason: 'Sports competition', status: 'pending', created_at: '2025-03-23' }},
];

// ---- Notifications ----
export const demoNotifications: Notification[] = [
  {{ id: 'n1', title: 'Welcome to Royal International Public School', message: 'Welcome to the academic year 2024-2025. We wish all students a great year!', recipient_role: undefined, created_by: 'u-admin', is_read: false, created_at: '2025-03-30T10:00:00' }},
  {{ id: 'n2', title: 'Attendance Alert', message: 'Low attendance detected for some students. Please ensure regular attendance.', recipient_role: 'parent', created_by: 'u-admin', is_read: false, created_at: '2025-03-29T09:30:00' }},
  {{ id: 'n3', title: 'Fee Reminder', message: 'Annual fee payment for the current session is due. Please clear pending dues.', recipient_role: 'parent', created_by: 'u-admin', is_read: false, created_at: '2025-03-28T14:00:00' }},
  {{ id: 'n4', title: 'Holiday Announcement', message: 'School will remain closed on April 14 for Ambedkar Jayanti.', recipient_role: undefined, created_by: 'u-admin', is_read: false, created_at: '2025-03-25T16:00:00' }},
  {{ id: 'n5', title: 'Parent-Teacher Meeting', message: 'PTM scheduled for April 5, 2025. All parents are requested to attend.', recipient_role: 'parent', created_by: 'u-admin', is_read: false, created_at: '2025-03-20T12:00:00' }},
];

// ---- Holidays ----
export const demoHolidays: Holiday[] = [
  {{ id: 'h1', title: 'Republic Day', holiday_date: '2025-01-26', description: 'National holiday', academic_year_id: 'ay1', created_at: '2024-01-01' }},
  {{ id: 'h2', title: 'Holi', holiday_date: '2025-03-14', description: 'Festival of colors', academic_year_id: 'ay1', created_at: '2024-01-01' }},
  {{ id: 'h3', title: 'Good Friday', holiday_date: '2025-04-18', description: 'Religious holiday', academic_year_id: 'ay1', created_at: '2024-01-01' }},
  {{ id: 'h4', title: 'Ambedkar Jayanti', holiday_date: '2025-04-14', description: 'National holiday', academic_year_id: 'ay1', created_at: '2024-01-01' }},
  {{ id: 'h5', title: 'Independence Day', holiday_date: '2025-08-15', description: 'National holiday', academic_year_id: 'ay1', created_at: '2024-01-01' }},
  {{ id: 'h6', title: 'Diwali', holiday_date: '2025-10-20', description: 'Festival of lights', academic_year_id: 'ay1', created_at: '2024-01-01' }},
  {{ id: 'h7', title: 'Christmas', holiday_date: '2025-12-25', description: 'Holiday season', academic_year_id: 'ay1', created_at: '2024-01-01' }},
];

// ---- School Settings ----
export const demoSchoolSettings: SchoolSettings = {{
  id: 'ss1',
  school_name: 'Royal International Public School',
  logo_url: '',
  minimum_attendance_percentage: 75,
  contact_email: '{school_email or "info@rips.edu.in"}',
  contact_phone: '{school_ph}',
  address: '{school_addr}',
  created_at: '2024-01-01',
  updated_at: '2024-01-01',
}};

// ---- Helper functions ----
export function getProfileById(id: string): Profile | undefined {{
  return demoProfiles.find(p => p.id === id);
}}

export function getStudentWithDetails(studentId: string) {{
  const student = demoStudents.find(s => s.id === studentId);
  if (!student) return null;
  return {{
    ...student,
    profile: getProfileById(student.profile_id),
    class: demoClasses.find(c => c.id === student.class_id),
    section: demoSections.find(s => s.id === student.section_id),
    parent: demoParents.find(p => p.id === student.parent_id),
  }};
}}

export function getTeacherWithProfile(teacherId: string) {{
  const teacher = demoTeachers.find(t => t.id === teacherId);
  if (!teacher) return null;
  return {{
    ...teacher,
    profile: getProfileById(teacher.profile_id),
  }};
}}

export function getStudentAttendanceStats(studentId: string) {{
  const records = demoAttendance.filter(a => a.student_id === studentId);
  const total = records.length;
  const present = records.filter(a => a.status === 'present').length;
  const absent = records.filter(a => a.status === 'absent').length;
  const late = records.filter(a => a.status === 'late').length;
  const excused = records.filter(a => a.status === 'excused').length;
  const percentage = total > 0 ? Math.round((present / total) * 100) : 0;
  return {{ total, present, absent, late, excused, percentage }};
}}

export function getClassAttendanceStats(classId: string, date?: string) {{
  let records = demoAttendance.filter(a => a.class_id === classId);
  if (date) records = records.filter(a => a.attendance_date === date);
  const total = records.length;
  const present = records.filter(a => a.status === 'present').length;
  const absent = records.filter(a => a.status === 'absent').length;
  const late = records.filter(a => a.status === 'late').length;
  return {{ total, present, absent, late, percentage: total > 0 ? Math.round((present / total) * 100) : 0 }};
}}

export function getTodayAttendanceSummary() {{
  const today = '2025-03-28';
  const records = demoAttendance.filter(a => a.attendance_date === today);
  const totalStudents = demoStudents.length;
  const present = records.filter(a => a.status === 'present').length;
  const absent = records.filter(a => a.status === 'absent').length;
  const late = records.filter(a => a.status === 'late').length;
  const excused = records.filter(a => a.status === 'excused').length;
  return {{
    totalStudents,
    present,
    absent,
    late,
    excused,
    percentage: totalStudents > 0 ? Math.round((present / totalStudents) * 100) : 0,
  }};
}}

export function getMonthlyTrend() {{
  const dates = [...new Set(demoAttendance.map(a => a.attendance_date))].sort();
  return dates.map(date => {{
    const records = demoAttendance.filter(a => a.attendance_date === date);
    const present = records.filter(a => a.status === 'present').length;
    const total = records.length;
    return {{
      date: date.slice(5),
      percentage: total > 0 ? Math.round((present / total) * 100) : 0,
      present,
      absent: records.filter(a => a.status === 'absent').length,
      late: records.filter(a => a.status === 'late').length,
    }};
  }});
}}

export function getClassWiseComparison() {{
  return demoClasses.map(cls => {{
    const stats = getClassAttendanceStats(cls.id);
    return {{ name: cls.name, ...stats }};
  }});
}}

export function getDefaulterList(threshold: number = 75) {{
  return demoStudents
    .slice(0, 50) // Only check students with attendance records
    .map(s => ({{
      ...s,
      profile: getProfileById(s.profile_id),
      class: demoClasses.find(c => c.id === s.class_id),
      section: demoSections.find(sec => sec.id === s.section_id),
      stats: getStudentAttendanceStats(s.id),
    }}))
    .filter(s => s.stats.percentage < threshold)
    .sort((a, b) => a.stats.percentage - b.stats.percentage);
}}
'''

with open('/Users/mohammad/Desktop/RPS/smart-school-attendance/src/lib/demo-data.ts', 'w') as f:
    f.write(output)

print(f"\n✅ Written demo-data.ts with {len(profiles)} profiles, {len(student_records)} students, {len(teacher_records)} teachers, {len(parent_records)} parents")
